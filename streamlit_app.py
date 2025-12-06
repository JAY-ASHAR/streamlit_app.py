# streamlit_app.py
# ---------------------------------
# Hostel Attendance Tracker with:
# - Admin & Operator login
# - Morning/Night attendance lock
# - Date-wise Excel report with totals
# - Admin-only student management
# - Admin-only Clear Data by Date
# - Dashboard-style totals on Generate Report
# - Edit after lock (Admin)
# - Analytics Dashboard (Bar Chart + Leaderboards)
# - Student Profiles with Excel export
# - Monthly Summary with Excel export
# - NEW: Leaderboards (Top Performers, Absentees, Consistency, Lowest Attendance)
# - NEW: Interactive Month + Student filters inside Analytics Dashboard
# ---------------------------------

import os
import json
from datetime import date, datetime
from typing import Dict, List
from io import BytesIO
import math

import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt

# ========== CONFIG ==========
APP_TITLE = "🏠 Hostel Attendance Tracker"
DATA_DIR = "data"
STUDENTS_CSV = os.path.join(DATA_DIR, "students.csv")
ATTENDANCE_DIR = os.path.join(DATA_DIR, "attendance")
LOCKS_JSON = os.path.join(DATA_DIR, "locks.json")
REPORTS_DIR = os.path.join(DATA_DIR, "reports")

SESSIONS = ["Morning", "Night"]
# statuses: Present, Absent, Leave, Sick, School/College, Office
STATUS_OPTIONS = ["P", "A", "L", "S", "SCH/CLG", "OI"]

USERS = {
    "warden1": {"password": "1234", "role": "admin", "name": "Warden 1"},
    "warden2": {"password": "1234", "role": "admin", "name": "Warden 2"},
    "morning": {"password": "1111", "role": "operator", "name": "Morning Operator"},
    "night": {"password": "2222", "role": "operator", "name": "Night Operator"},
}

# ========== HELPERS ==========
def ensure_dirs():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(ATTENDANCE_DIR, exist_ok=True)
    os.makedirs(REPORTS_DIR, exist_ok=True)


def ensure_students_csv():
    if not os.path.exists(STUDENTS_CSV):
        df = pd.DataFrame([
            {"student_id": 1, "name": "Student One", "active": True},
            {"student_id": 2, "name": "Student Two", "active": True},
        ])
        df.to_csv(STUDENTS_CSV, index=False)


def load_students() -> pd.DataFrame:
    ensure_students_csv()
    df = pd.read_csv(STUDENTS_CSV)
    if "active" not in df.columns:
        df["active"] = True
    return df[df["active"] == True].copy()


def save_students(df: pd.DataFrame):
    df.to_csv(STUDENTS_CSV, index=False)


def locks_read() -> Dict:
    if not os.path.exists(LOCKS_JSON):
        with open(LOCKS_JSON, "w") as f:
            json.dump({}, f)
    with open(LOCKS_JSON, "r") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {}


def locks_write(data: Dict):
    with open(LOCKS_JSON, "w") as f:
        json.dump(data, f, indent=2)


def is_session_locked(day_str: str, session: str) -> bool:
    locks = locks_read()
    return locks.get(day_str, {}).get(session, False)


def lock_session(day_str: str, session: str):
    locks = locks_read()
    if day_str not in locks:
        locks[day_str] = {}
    locks[day_str][session] = True
    locks_write(locks)


def unlock_session(day_str: str, session: str):
    locks = locks_read()
    if day_str in locks and session in locks[day_str]:
        locks[day_str][session] = False
        locks_write(locks)


def daily_csv_path(day_str: str) -> str:
    return os.path.join(ATTENDANCE_DIR, f"{day_str}.csv")


def load_or_init_daily(day_str: str) -> pd.DataFrame:
    path = daily_csv_path(day_str)
    active_students = load_students()[["student_id", "name"]].copy()
    if os.path.exists(path):
        df = pd.read_csv(path)
        # merge so new students appear and ordering matches students file
        df = active_students.merge(df, on=["student_id", "name"], how="left")
    else:
        df = active_students.copy()
        df["Morning"] = ""
        df["Night"] = ""
    # ensure columns exist
    for col in ["Morning", "Night"]:
        if col not in df.columns:
            df[col] = ""
    return df


def save_daily(day_str: str, df: pd.DataFrame):
    df.to_csv(daily_csv_path(day_str), index=False)


def session_totals(df: pd.DataFrame, session: str) -> Dict[str, int]:
    counts = df[session].fillna("").value_counts().to_dict()
    return {k: int(counts.get(k, 0)) for k in STATUS_OPTIONS}


def create_excel_with_totals(df: pd.DataFrame, session: str) -> BytesIO:
    totals = session_totals(df, session)
    output = BytesIO()
    df_out = df[["name", session]].copy()
    df_out.columns = ["Name", session]

    summary_df = pd.DataFrame({
        "Metric": [f"Total {s}" for s in STATUS_OPTIONS],
        "Count": [totals[s] for s in STATUS_OPTIONS]
    })

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Attendance", startrow=0)
        summary_df.to_excel(writer, index=False, sheet_name="Attendance", startrow=len(df_out) + 2)
    output.seek(0)
    return output


def create_combined_excel(df: pd.DataFrame) -> BytesIO:
    output = BytesIO()
    df_out = df[["name", "Morning", "Night"]].copy()
    df_out.columns = ["Name", "Morning", "Night"]

    rows = []
    for sess in SESSIONS:
        totals = session_totals(df, sess)
        for s in STATUS_OPTIONS:
            rows.append({"Metric": f"{sess} {s}", "Count": totals[s]})
    summary_df = pd.DataFrame(rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Attendance", startrow=0)
        summary_df.to_excel(writer, index=False, sheet_name="Attendance", startrow=len(df_out) + 2)
    output.seek(0)
    return output


def clear_data_by_date(day_str: str):
    att_file = daily_csv_path(day_str)
    if os.path.exists(att_file):
        os.remove(att_file)
    for f in os.listdir(REPORTS_DIR):
        if day_str in f:
            try:
                os.remove(os.path.join(REPORTS_DIR, f))
            except Exception:
                pass
    locks = locks_read()
    if day_str in locks:
        del locks[day_str]
        locks_write(locks)

# ========== ANALYTICS HELPERS (NEW) ==========
@st.cache_data(show_spinner=False)
def load_all_attendance_long() -> pd.DataFrame:
    """Read all daily CSVs and return a long dataframe:
    columns: date(str YYYY-MM-DD), name, student_id, session, status
    """
    if not os.path.exists(ATTENDANCE_DIR):
        return pd.DataFrame(columns=["date", "name", "student_id", "session", "status"])    
    files = sorted([f for f in os.listdir(ATTENDANCE_DIR) if f.endswith(".csv")])
    rows: List[pd.DataFrame] = []
    for f in files:
        try:
            df = pd.read_csv(os.path.join(ATTENDANCE_DIR, f))
            if "student_id" not in df.columns:
                # fall back in case of legacy files
                df = load_students()[["student_id", "name"]].merge(df, on=["name"], how="right")
            day = f.replace(".csv", "")
            df["date"] = day
            part = df.melt(id_vars=["student_id", "name", "date"], value_vars=SESSIONS,
                           var_name="session", value_name="status")
            rows.append(part)
        except Exception:
            # skip bad files but keep app running
            continue
    if not rows:
        return pd.DataFrame(columns=["date", "name", "student_id", "session", "status"])    
    out = pd.concat(rows, ignore_index=True)
    # normalize blanks
    out["status"] = out["status"].fillna("")
    # add datetime + month
    out["date_dt"] = pd.to_datetime(out["date"], errors="coerce")
    out["month"] = out["date_dt"].dt.strftime("%Y-%m")
    return out


def compute_student_stats(df_long: pd.DataFrame) -> pd.DataFrame:
    """Return per-student aggregate stats over df_long selection.
    - total_sessions
    - present (P) count
    - absent (A) count
    - attendance_rate = present / total_sessions
    - daily_consistency_std: std dev of daily present ratio (0, 0.5, 1.0)
    """
    if df_long.empty:
        return pd.DataFrame(columns=[
            "student_id","name","total_sessions","P","A","L","S","SCH/CLG","OI",
            "attendance_rate","days_covered","daily_consistency_std"
        ])

    # per student totals by status
    pivot = (df_long
             .groupby(["student_id", "name", "status"], dropna=False)
             .size()
             .unstack(fill_value=0))

    # make sure all status columns exist
    for s in STATUS_OPTIONS:
        if s not in pivot.columns:
            pivot[s] = 0

    pivot = pivot.reset_index()
    pivot["total_sessions"] = pivot[STATUS_OPTIONS].sum(axis=1)
    pivot["P"] = pivot["P"].astype(int)
    pivot["A"] = pivot["A"].astype(int)

    # attendance rate
    pivot["attendance_rate"] = np.where(
        pivot["total_sessions"] > 0,
        pivot["P"] / pivot["total_sessions"],
        np.nan
    )

    # daily present ratio series for consistency metric
    # For each student and day, present_ratio = (#P on that day across sessions)/2
    day_level = (df_long.assign(present=lambda x: (x["status"] == "P").astype(int))
                 .groupby(["student_id", "name", "date"], as_index=False)["present"].sum())
    day_level["daily_ratio"] = day_level["present"] / len(SESSIONS)

    # std dev per student (NaN if only one day)
    stds = (day_level.groupby(["student_id", "name"])['daily_ratio']
            .agg(lambda s: float(np.std(s, ddof=0)) if len(s) > 0 else np.nan)
            .reset_index()
            .rename(columns={"daily_ratio": "daily_consistency_std"}))

    days = (day_level.groupby(["student_id", "name"])['date']
            .nunique().reset_index().rename(columns={"date": "days_covered"}))

    out = (pivot
           .merge(stds, on=["student_id", "name"], how="left")
           .merge(days, on=["student_id", "name"], how="left"))

    # order columns
    cols = ["student_id", "name", "total_sessions"] + STATUS_OPTIONS + [
        "attendance_rate", "days_covered", "daily_consistency_std"
    ]
    out = out[cols]
    return out


# ========== NEW FEATURES ==========
def analytics_dashboard():
    st.header("📊 Attendance Analytics (Admin Only)")

    # Load data
    df_long = load_all_attendance_long()
    if df_long.empty:
        st.info("No attendance data available.")
        return

    # ----- Filters (Month + Student) -----
    months = sorted(df_long["month"].dropna().unique().tolist())
    selected_month = st.selectbox("Select Month", ["All"] + months, index=0)

    df_filtered = df_long.copy()
    if selected_month != "All":
        df_filtered = df_filtered[df_filtered["month"] == selected_month]

    # student list from filtered
    student_names = sorted(df_filtered["name"].dropna().unique().tolist())
    selected_students = st.multiselect("Filter Students (optional)", student_names, default=student_names)

    if selected_students:
        df_filtered = df_filtered[df_filtered["name"].isin(selected_students)]

    # Session selector for bar summary
    session_for_bar = st.selectbox("Select Session for Summary Bar Chart", SESSIONS, index=0)

    # Summary counts bar (by status)
    counts = (df_filtered[df_filtered["session"] == session_for_bar]["status"].fillna("")
              .value_counts().reindex(STATUS_OPTIONS, fill_value=0))

    fig, ax = plt.subplots()
    counts.plot(kind="bar", ax=ax)
    ax.set_title(f"{session_for_bar} Attendance Summary ({'All months' if selected_month=='All' else selected_month})")
    ax.set_xlabel("Status")
    ax.set_ylabel("Count")
    st.pyplot(fig, use_container_width=True)

    st.divider()

    # ----- Leaderboards -----
    st.subheader("🏆 Leaderboards")

    stats = compute_student_stats(df_filtered)

    # Minimum days filter to avoid tiny-sample noise
    min_days = st.slider("Minimum days to qualify", min_value=1, max_value=int(stats["days_covered"].max() if not stats.empty else 1), value=3)
    eligible = stats[stats["days_covered"] >= min_days].copy()

    if eligible.empty:
        st.info("No students meet the selected criteria yet.")
    else:
        # Top Performers (highest attendance rate)
        top_performers = (eligible.sort_values(["attendance_rate", "days_covered", "P"], ascending=[False, False, False])
                                   .head(10)
                                   .reset_index(drop=True))

        # Absentees (highest total A)
        top_absentees = (eligible.sort_values(["A", "attendance_rate"], ascending=[False, True])
                                  .head(10)
                                  .reset_index(drop=True))

        # Most Consistent (lowest std of daily present ratio); tie-breaker: higher attendance_rate
        consistent = (eligible.sort_values(["daily_consistency_std", "attendance_rate"], ascending=[True, False])
                               .head(10)
                               .reset_index(drop=True))

        # Lowest Attendance (lowest attendance rate)
        lowest_attendance = (eligible.sort_values(["attendance_rate", "days_covered"], ascending=[True, False])
                                      .head(10)
                                      .reset_index(drop=True))

        # ----- Display Leaderboards -----
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Top Performers (by Attendance Rate)**")
            tmp = top_performers.copy()
            tmp["attendance_rate"] = (tmp["attendance_rate"] * 100).round(2).astype(float)
            tmp = tmp.rename(columns={"attendance_rate": "Attendance %"})
            show_cols = ["name", "Attendance %", "P", "A", "days_covered", "total_sessions"]
            st.dataframe(tmp[show_cols], use_container_width=True)

        with col2:
            st.markdown("**Top Absentees (by Total A)**")
            tmp = top_absentees.copy()
            tmp["attendance_rate"] = (tmp["attendance_rate"] * 100).round(2).astype(float)
            tmp = tmp.rename(columns={"attendance_rate": "Attendance %"})
            show_cols = ["name", "A", "P", "days_covered", "Attendance %"]
            st.dataframe(tmp[show_cols], use_container_width=True)

        col3, col4 = st.columns(2)
        with col3:
            st.markdown("**Most Consistent (lowest day-to-day variation)**")
            tmp = consistent.copy()
            tmp["attendance_rate"] = (tmp["attendance_rate"] * 100).round(2).astype(float)
            tmp["daily_consistency_std"] = tmp["daily_consistency_std"].round(3)
            tmp = tmp.rename(columns={"attendance_rate": "Attendance %", "daily_consistency_std": "Consistency Std"})
            show_cols = ["name", "Consistency Std", "Attendance %", "days_covered", "total_sessions"]
            st.dataframe(tmp[show_cols], use_container_width=True)

        with col4:
            st.markdown("**Lowest Attendance (by Attendance Rate)**")
            tmp = lowest_attendance.copy()
            tmp["attendance_rate"] = (tmp["attendance_rate"] * 100).round(2).astype(float)
            tmp = tmp.rename(columns={"attendance_rate": "Attendance %"})
            show_cols = ["name", "Attendance %", "P", "A", "days_covered", "total_sessions"]
            st.dataframe(tmp[show_cols], use_container_width=True)

    st.caption("Notes: Attendance % = P / total sessions. Consistency uses std dev of daily present ratio (0, 0.5, 1.0).")

# ========== UI PAGES ==========
def login_ui():
    st.title(APP_TITLE)
    st.subheader("Login")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Sign in", type="primary"):
        user = USERS.get(username)
        if user and user.get("password") == password:
            st.session_state["user"] = {**user, "username": username}
            st.success(f"Welcome, {user['name']}!")
            st.rerun()
        else:
            st.error("Invalid username or password.")


def nav_sidebar():
    user = st.session_state.get("user")
    st.sidebar.title("Navigation")
    st.sidebar.write(f"👤 Logged in as: **{user['name']}** ({user['role']})")
    options = ["Take Attendance", "Generate Report"]
    if user["role"] == "admin":
        options += ["Manage Students", "Clear Data", "Analytics Dashboard", "Student Profiles"]
    choice = st.sidebar.radio("Go to", options, index=0)
    st.sidebar.divider()
    if st.sidebar.button("Logout"):
        st.session_state.clear()
        st.rerun()
    return choice


def take_attendance_page():
    st.header("📝 Take Attendance")
    day = st.date_input("Select date", value=date.today())
    day_str = day.strftime("%Y-%m-%d")

    user = st.session_state["user"]
    if user["role"] == "admin":
        allowed_sessions = SESSIONS
    else:
        allowed_sessions = ["Morning"] if user["username"] == "morning" else ["Night"]

    session = st.selectbox("Session", allowed_sessions, index=0)

    locked = is_session_locked(day_str, session)
    df = load_or_init_daily(day_str)

    if locked and user["role"] == "admin":
        st.warning(f"{session} attendance for {day_str} is locked. Admin can unlock/edit.")
        if st.button("Unlock for Admin Edit"):
            unlock_session(day_str, session)
            st.success("Unlocked. You can now edit and re-lock after changes.")
            st.rerun()
    elif locked:
        st.info(f"✅ {session} attendance for {day_str} is locked. No changes allowed.")
        st.dataframe(df[["name", "Morning", "Night"]].rename(columns={"name": "Name"}), use_container_width=True)
        return

    st.caption("Mark each student as one of the statuses. Admins can edit locked sessions (unlock).")

    updates = []
    for i, row in df.iterrows():
        col1, col2 = st.columns([3, 2])
        with col1:
            st.write(row["name"])
        with col2:
            current = row[session] if pd.notna(row[session]) and (row[session] in STATUS_OPTIONS) else ""
            choice = st.radio(
                label=f"status_{day_str}_{session}_{i}",
                options=["", *STATUS_OPTIONS],
                index=["", *STATUS_OPTIONS].index(current) if current in ["", *STATUS_OPTIONS] else 0,
                horizontal=True,
                label_visibility="collapsed",
            )
        updates.append(choice)

    df[session] = updates

    st.divider()
    if st.button(f"Submit & Lock {session} Attendance", type="primary"):
        if any(x == "" for x in updates):
            st.warning("Please mark all students before submitting.")
        else:
            save_daily(day_str, df)
            lock_session(day_str, session)
            st.success(f"{session} attendance locked for {day_str}.")
            st.rerun()

    totals = session_totals(df, session)
    totals_str = " | ".join([f"{k}: {totals[k]}" for k in STATUS_OPTIONS])
    st.markdown(f"**Totals ({session})** — {totals_str}")


def generate_report_page():
    st.header("📊 Generate Excel Report")
    day = st.date_input("Select date", value=date.today(), key="report_date")
    day_str = day.strftime("%Y-%m-%d")

    user = st.session_state["user"]
    if user["role"] == "admin":
        allowed_sessions = SESSIONS
    else:
        allowed_sessions = ["Morning"] if user["username"] == "morning" else ["Night"]

    if not os.path.exists(daily_csv_path(day_str)):
        st.error("No attendance data file for this date.")
        return

    df = pd.read_csv(daily_csv_path(day_str))

    # show lock status and dashboard-style small cards
    m_locked = is_session_locked(day_str, "Morning")
    n_locked = is_session_locked(day_str, "Night")
    st.write(f"Morning locked: {'✅' if m_locked else '❌'} | Night locked: {'✅' if n_locked else '❌'}")

    st.subheader("Totals (this date)")
    col0, col1 = st.columns(2)
    with col0:
        st.markdown("**Morning**")
        totals_m = session_totals(df, "Morning")
        cols = st.columns(len(STATUS_OPTIONS))
        for c, s in zip(cols, STATUS_OPTIONS):
            c.metric(label=s, value=str(totals_m[s]))
    with col1:
        st.markdown("**Night**")
        totals_n = session_totals(df, "Night")
        cols2 = st.columns(len(STATUS_OPTIONS))
        for c, s in zip(cols2, STATUS_OPTIONS):
            c.metric(label=s, value=str(totals_n[s]))

    # Downloads per allowed session
    for sess in allowed_sessions:
        st.subheader(f"{sess} Report")
        file_obj = create_excel_with_totals(df, sess)
        st.download_button(
            label=f"Download {sess} Report (.xlsx)",
            data=file_obj,
            file_name=f"{sess.lower()}_{day_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Admin-only combined report (table + summary)
    if user["role"] == "admin":
        st.subheader("Combined Attendance (single-sheet) — Admin Export")
        combined = create_combined_excel(df)
        st.download_button(
            label="Download Combined Attendance (.xlsx)",
            data=combined,
            file_name=f"attendance_combined_{day_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def manage_students_page():
    st.header("👥 Manage Students (Admins Only)")
    df = pd.read_csv(STUDENTS_CSV)

    st.subheader("Add Student")
    new_id = st.text_input("Enter Student ID", key="add_id")
    new_name = st.text_input("Full name", key="add_name")
    if st.button("Add Student"):
        if not new_id.strip().isdigit() or new_name.strip() == "":
            st.warning("Enter valid numeric ID and name.")
        else:
            if int(new_id) in df["student_id"].tolist():
                st.error("Student ID already exists.")
            else:
                df.loc[len(df)] = {"student_id": int(new_id), "name": new_name.strip(), "active": True}
                save_students(df)
                st.success(f"Added: {new_name.strip()}")
                st.rerun()

    st.subheader("Active Students")
    active_df = df[df["active"] == True].copy()
    st.dataframe(active_df[["student_id", "name"]], use_container_width=True)

    st.subheader("Rename / Deactivate Student")
    if not active_df.empty:
        sel = st.selectbox("Select student", active_df["name"].tolist(), key="manage_sel")
        col1, col2 = st.columns(2)
        with col1:
            new_label = st.text_input("Rename to", value=sel, key="rename_box")
            if st.button("Rename"):
                if new_label.strip():
                    df.loc[df["name"] == sel, "name"] = new_label.strip()
                    save_students(df)
                    st.success(f"Renamed to: {new_label.strip()}")
                    st.rerun()
        with col2:
            if st.button("Deactivate / Delete"):
                df.loc[df["name"] == sel, "active"] = False
                save_students(df)
                st.success(f"Deactivated: {sel}")
                st.rerun()
    else:
        st.info("No active students available.")


def clear_data_page():
    st.header("🗑 Clear Data by Date (Admins Only)")
    day = st.date_input("Select date to clear", value=date.today(), key="clear_date")
    day_str = day.strftime("%Y-%m-%d")
    st.write("This will remove the attendance file, any generated reports with this date in the file name, and stored locks for that date.")
    if st.button("Clear Data for Selected Date (Admin only)", type="secondary"):
        clear_data_by_date(day_str)
        st.success(f"All data for {day_str} has been cleared.")
        st.rerun()


# ========== STUDENT PROFILES ==========
def student_profiles():
    st.header("👤 Student Profiles")

    # Load students
    df_students = load_students()
    if df_students.empty:
        st.info("No students available.")
        return

    # Select student
    student_list = df_students["name"].tolist()
    selected_student = st.selectbox("Select Student", student_list)

    # Load all attendance (long format)
    df_long = load_all_attendance_long()
    if df_long.empty:
        st.warning("No attendance data available yet.")
        return

    # Filter for selected student
    df_individual = df_long[df_long["name"] == selected_student].copy()
    df_individual = df_individual.sort_values(["date_dt", "session"])

    # Pivot so Morning & Night are in separate columns
    df_pivot = df_individual.pivot_table(
        index="date", columns="session", values="status", aggfunc="first"
    ).reset_index()

    # Ensure both Morning & Night columns exist
    for col in SESSIONS:
        if col not in df_pivot.columns:
            df_pivot[col] = ""

    # Show in Streamlit
    st.subheader(f"📅 Attendance History — {selected_student}")
    st.dataframe(df_pivot[["date", "Morning", "Night"]], use_container_width=True)

    # Prepare Excel file with colors
    if not df_pivot.empty:
        from io import BytesIO
        from openpyxl.styles import PatternFill

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_pivot[["date", "Morning", "Night"]].to_excel(
                writer, sheet_name="Attendance", index=False
            )

            # Access the sheet
            worksheet = writer.sheets["Attendance"]

            # Define fills
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Loop through cells and apply colors
            for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=3):
                for cell in row:
                    if cell.value in ["P", "L", "S", "SCH/CLG", "OI"]:
                        cell.fill = green_fill
                    elif cell.value == "A":
                        cell.fill = red_fill

        output.seek(0)

        # Download button
        st.download_button(
            label=f"Download {selected_student}'s Report (Excel)",
            data=output,
            file_name=f"{selected_student.replace(' ', '_')}_attendance.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ========== MAIN ==========
def main():
    st.set_page_config(page_title=APP_TITLE, page_icon="🏠", layout="wide")
    ensure_dirs()
    ensure_students_csv()

    if "user" not in st.session_state:
        login_ui()
        return

    choice = nav_sidebar()
    if choice == "Take Attendance":
        take_attendance_page()
    elif choice == "Generate Report":
        generate_report_page()
    elif choice == "Manage Students":
        if st.session_state["user"]["role"] != "admin":
            st.error("Access denied.")
            return
        manage_students_page()
    elif choice == "Clear Data":
        if st.session_state["user"]["role"] != "admin":
            st.error("Access denied.")
            return
        clear_data_page()
    elif choice == "Analytics Dashboard":
        if st.session_state["user"]["role"] != "admin":
            st.error("Access denied.")
            return
        analytics_dashboard()
    elif choice == "Student Profiles":
        if st.session_state["user"]["role"] != "admin":
            st.error("Access denied.")
            return
        student_profiles()

if __name__ == "__main__":
    main()
